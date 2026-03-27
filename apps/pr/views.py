"""
PR application views for Purchase Requisition CRUD operations,
approval workflow, and status tracking.
"""

from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView
)
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from django.urls import reverse_lazy
from django.db.models import Q, Sum, F
from django.utils import timezone
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST
from django.utils.decorators import method_decorator
import json

from .models import PurchaseRequisition, PRItem, ApprovalChain, PRApproval
from .forms import (
    PRHeaderForm, PRItemForm, PRItemFormSet, ApprovalForm, 
    PRFilterForm, BulkActionForm, VendorQuotationForm
)
from .services import PRApprovalService, PRGenerationService
from apps.users.permissions import require_permission
from apps.audit.models import AuditLog
from apps.attachments.models import Attachment


class PRListView(LoginRequiredMixin, ListView):
    """Display list of all PRs with filtering and pagination."""
    
    model = PurchaseRequisition
    template_name = 'pr/pr_list.html'
    context_object_name = 'prs'
    paginate_by = 20
    
    def get_queryset(self):
        """Filter PRs based on user role and search criteria."""
        queryset = PurchaseRequisition.objects.select_related(
            'created_by', 'last_modified_by'
        ).prefetch_related('approvals')
        
        # Filter by user role
        user = self.request.user
        if user.role.name == 'Requester':
            queryset = queryset.filter(created_by=user)
        elif user.role.name == 'Approver':
            # Show PRs pending approval for this user
            queryset = queryset.filter(
                Q(approvals__approver=user, approvals__approval_status='PENDING')
            ).distinct()
        
        # Apply search filters
        form = PRFilterForm(self.request.GET)
        if form.is_valid():
            search = form.cleaned_data.get('search')
            if search:
                queryset = queryset.filter(
                    Q(pr_id__icontains=search) |
                    Q(requester_name__icontains=search) |
                    Q(purpose_of_requirement__icontains=search)
                )
            
            status = form.cleaned_data.get('status')
            if status:
                queryset = queryset.filter(status=status)
            
            pr_type = form.cleaned_data.get('pr_type')
            if pr_type:
                queryset = queryset.filter(pr_type=pr_type)
            
            date_from = form.cleaned_data.get('date_from')
            if date_from:
                queryset = queryset.filter(pr_date__gte=date_from)
            
            date_to = form.cleaned_data.get('date_to')
            if date_to:
                queryset = queryset.filter(pr_date__lte=date_to)
        
        return queryset.order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = PRFilterForm(self.request.GET)
        
        # Add summary statistics
        queryset = self.get_queryset()
        context['total_prs'] = queryset.count()
        context['pending_prs'] = queryset.filter(status='PENDING_APPROVAL').count()
        context['approved_prs'] = queryset.filter(status='APPROVED').count()
        context['total_value'] = queryset.aggregate(
            total=Sum('items__total_value')
        )['total'] or 0
        
        return context


class PRDetailView(LoginRequiredMixin, DetailView):
    """Display detailed view of a single PR."""
    
    model = PurchaseRequisition
    template_name = 'pr/pr_detail.html'
    context_object_name = 'pr'
    slug_field = 'pr_id'
    slug_url_kwarg = 'pr_id'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pr = self.get_object()
        
        context['items'] = pr.items.all()
        context['approvals'] = pr.approvals.select_related('approver')
        context['attachments'] = pr.attachments.all()
        context['activity_logs'] = pr.activity_logs.all().order_by('-created_at')[:10]
        
        # Check if current user can approve
        user = self.request.user
        pending_approval = pr.approvals.filter(
            approver=user, approval_status='PENDING'
        ).first()
        
        if pending_approval:
            context['approval_form'] = ApprovalForm()
            context['can_approve'] = True
        
        return context


class PRCreateView(LoginRequiredMixin, CreateView):
    """Create a new PR."""
    
    model = PurchaseRequisition
    form_class = PRHeaderForm
    template_name = 'pr/pr_form.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if self.request.POST:
            context['formset'] = PRItemFormSet(self.request.POST)
        else:
            context['formset'] = PRItemFormSet()
        
        return context
    
    def form_valid(self, form):
        """Handle form submission and create PR with items."""
        context = self.get_context_data()
        formset = context['formset']
        
        if formset.is_valid():
            # Create PR
            pr = form.save(commit=False)
            pr.created_by = self.request.user
            pr.status = 'DRAFT'
            pr.save()
            
            # Create items
            for item_form in formset:
                if item_form.cleaned_data:
                    item = item_form.save(commit=False)
                    item.pr = pr
                    item.save()
            
            # Determine approval requirements
            total_value = sum(
                item.total_value or 0 for item in pr.items.all()
            )
            
            # Create approval chain based on PR value and type
            approval_service = PRApprovalService()
            approval_service.create_approval_chain(pr, total_value)
            
            messages.success(
                self.request,
                f'PR {pr.pr_id} created successfully.'
            )
            return redirect('pr:pr-detail', pr_id=pr.pr_id)
        else:
            return self.form_invalid(form)
    
    def get_success_url(self):
        return reverse_lazy('pr:pr-detail', kwargs={'pr_id': self.object.pr_id})


class PRUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Edit an existing PR (only if in DRAFT status)."""
    
    model = PurchaseRequisition
    form_class = PRHeaderForm
    template_name = 'pr/pr_form.html'
    slug_field = 'pr_id'
    slug_url_kwarg = 'pr_id'
    
    def test_func(self):
        """Only creator can edit, and only in DRAFT status."""
        pr = self.get_object()
        return (
            pr.created_by == self.request.user and
            pr.status == 'DRAFT'
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if self.request.POST:
            context['formset'] = PRItemFormSet(self.request.POST, instance=self.object)
        else:
            context['formset'] = PRItemFormSet(instance=self.object)
        
        return context
    
    def form_valid(self, form):
        """Update PR and items."""
        context = self.get_context_data()
        formset = context['formset']
        
        if formset.is_valid():
            pr = form.save(commit=False)
            pr.last_modified_by = self.request.user
            pr.save()
            
            formset.save()
            
            messages.success(self.request, 'PR updated successfully.')
            return redirect('pr:pr-detail', pr_id=pr.pr_id)
        else:
            return self.form_invalid(form)


class PRSubmitView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Submit PR for approval."""
    
    model = PurchaseRequisition
    fields = []
    slug_field = 'pr_id'
    slug_url_kwarg = 'pr_id'
    
    def test_func(self):
        """Only creator can submit."""
        pr = self.get_object()
        return (
            pr.created_by == self.request.user and
            pr.status == 'DRAFT'
        )
    
    def post(self, request, *args, **kwargs):
        pr = self.get_object()
        
        # Validate PR before submission
        if not pr.items.exists():
            messages.error(request, 'PR must have at least one item.')
            return redirect('pr:pr-detail', pr_id=pr.pr_id)
        
        # Check all items have required fields
        for item in pr.items.all():
            if not item.short_text or not item.quantity or not item.unit_price:
                messages.error(request, 'All items must have text, quantity, and price.')
                return redirect('pr:pr-detail', pr_id=pr.pr_id)
        
        # Update status to pending approval
        pr.status = 'PENDING_APPROVAL'
        pr.submitted_date = timezone.now()
        pr.save()
        
        # Log action
        AuditLog.objects.create(
            user=request.user,
            action='SUBMITTED_FOR_APPROVAL',
            object_type='PurchaseRequisition',
            object_id=pr.id,
            description=f'PR {pr.pr_id} submitted for approval'
        )
        
        messages.success(request, f'PR {pr.pr_id} submitted for approval.')
        return redirect('pr:pr-detail', pr_id=pr.pr_id)


@require_POST
def approve_pr(request, pr_id):
    """Approve a PR."""
    pr = get_object_or_404(PurchaseRequisition, pr_id=pr_id)
    
    # Check if user is an approver
    pending_approval = pr.approvals.filter(
        approver=request.user, approval_status='PENDING'
    ).first()
    
    if not pending_approval:
        messages.error(request, 'You are not authorized to approve this PR.')
        return redirect('pr:pr-detail', pr_id=pr_id)
    
    form = ApprovalForm(request.POST, instance=pending_approval)
    
    if form.is_valid():
        approval = form.save(commit=False)
        approval.approval_date = timezone.now()
        approval.save()
        
        # Check if all approvals are complete
        approval_service = PRApprovalService()
        pr.status = approval_service.update_pr_status(pr)
        pr.save()
        
        # Log action
        action = 'APPROVED' if approval.approval_status == 'APPROVED' else 'REJECTED'
        AuditLog.objects.create(
            user=request.user,
            action=action,
            object_type='PurchaseRequisition',
            object_id=pr.id,
            description=f'PR {pr.pr_id} {action.lower()} by {request.user.get_full_name()}'
        )
        
        messages.success(request, f'PR {action.lower()} successfully.')
    
    return redirect('pr:pr-detail', pr_id=pr_id)


class MyPendingApprovalsView(LoginRequiredMixin, ListView):
    """View PRs pending approval for current user."""
    
    template_name = 'pr/pending_approvals.html'
    context_object_name = 'pending_approvals'
    paginate_by = 20
    
    def get_queryset(self):
        """Get PRs awaiting approval from current user."""
        return PRApproval.objects.filter(
            approver=self.request.user,
            approval_status='PENDING'
        ).select_related('purchase_requisition').order_by('-created_at')


class PRExportView(LoginRequiredMixin, DetailView):
    """Export PR details to Excel."""
    
    model = PurchaseRequisition
    slug_field = 'pr_id'
    slug_url_kwarg = 'pr_id'
    
    def get(self, request, *args, **kwargs):
        pr = self.get_object()
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messages.error(request, 'Excel export requires openpyxl package.')
            return redirect('pr:pr-detail', pr_id=pr.pr_id)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'PR Details'
        
        # Add headers and PR info
        ws['A1'] = f'Purchase Requisition {pr.pr_id}'
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        ws[f'A{row}'] = 'Requester:'
        ws[f'B{row}'] = pr.requester_name
        
        row += 1
        ws[f'A{row}'] = 'Department:'
        ws[f'B{row}'] = pr.department
        
        row += 1
        ws[f'A{row}'] = 'Purpose:'
        ws[f'B{row}'] = pr.purpose_of_requirement
        
        row += 1
        ws[f'A{row}'] = 'Date:'
        ws[f'B{row}'] = pr.pr_date.strftime('%Y-%m-%d')
        
        row += 2
        
        # Add items table
        headers = ['Item #', 'Description', 'Qty', 'Unit', 'Unit Price', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        row += 1
        for item in pr.items.all():
            ws.cell(row=row, column=1).value = item.item_number
            ws.cell(row=row, column=2).value = item.short_text
            ws.cell(row=row, column=3).value = float(item.quantity)
            ws.cell(row=row, column=4).value = item.unit
            ws.cell(row=row, column=5).value = float(item.unit_price or 0)
            ws.cell(row=row, column=6).value = float(item.total_value or 0)
            row += 1
        
        # Add total
        ws.cell(row=row, column=5).value = 'Total:'
        ws.cell(row=row, column=5).font = Font(bold=True)
        ws.cell(row=row, column=6).value = float(
            pr.items.aggregate(Sum('total_value'))['total_value__sum'] or 0
        )
        ws.cell(row=row, column=6).font = Font(bold=True)
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # Return file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="PR_{pr.pr_id}.xlsx"'
        wb.save(response)
        return response
